﻿# -*- coding: utf-8 -*-
"""
/***************************************************************************
 LDMP - A QGIS plugin
 This plugin supports monitoring and reporting of land degradation to the UNCCD 
 and in support of the SDG Land Degradation Neutrality (LDN) target.
                              -------------------
        begin                : 2017-05-23
        git sha              : $Format:%H$
        copyright            : (C) 2017 by Conservation International
        email                : trends.earth@conservation.org
 ***************************************************************************/
"""

from builtins import range
import os
import json
import tempfile

import numpy as np

from osgeo import ogr, osr, gdal

import openpyxl
from openpyxl.drawing.image import Image

from qgis.PyQt import QtWidgets, uic, QtXml
from qgis.PyQt.QtCore import QSettings

from qgis.core import QgsGeometry
from qgis.utils import iface
mb = iface.messageBar()

from LDMP import log
from LDMP.api import run_script
from LDMP.calculate import DlgCalculateBase, get_script_slug, ClipWorker
from LDMP.lc_setup import lc_setup_widget, lc_define_deg_widget
from LDMP.layers import add_layer, create_local_json_metadata, get_band_infos
from LDMP.schemas.schemas import BandInfo, BandInfoSchema
from LDMP.gui.DlgCalculateOneStep import Ui_DlgCalculateOneStep
from LDMP.gui.DlgCalculateSummaryTableAdmin import Ui_DlgCalculateSummaryTableAdmin
from LDMP.worker import AbstractWorker, StartWorker
from LDMP.summary import *


class DlgCalculateOneStep(DlgCalculateBase, Ui_DlgCalculateOneStep):
    def __init__(self, parent=None):
        super(DlgCalculateOneStep, self).__init__(parent)

        self.setupUi(self)

        # TODO:
        # Set max/min values for year_start and year_final, taking into account 
        # the ESA and MODIS data availability years
        
    def showEvent(self, event):
        super(DlgCalculateOneStep, self).showEvent(event)

        self.lc_setup_tab = lc_setup_widget
        self.TabBox.insertTab(1, self.lc_setup_tab, self.tr('Land Cover Setup'))

        # TODO: Temporarily hide these boxes until custom LC support for SOC is 
        # implemented
        self.lc_setup_tab.use_esa.setChecked(True)
        self.lc_setup_tab.use_custom.hide()
        self.lc_setup_tab.groupBox_custom_bl.hide()
        self.lc_setup_tab.groupBox_custom_tg.hide()

        self.lc_define_deg_tab = lc_define_deg_widget
        self.TabBox.insertTab(2, self.lc_define_deg_tab, self.tr('Define Effects of Land Cover Change'))
        
        # Hide the land cover ESA period box, since only one period is used in 
        # this dialog - the one on the main setup tab
        self.lc_setup_tab.groupBox_esa_period.hide()

        #######################################################################
        #######################################################################
        # Hack to calculate multiple countries at once for workshop preparation
        #######################################################################
        #######################################################################
        # from PyQt4.QtCore import QTimer, Qt
        # from PyQt4.QtWidgets import QMessageBox, QApplication
        # from PyQt4.QtTest import QTest
        # from LDMP.download import read_json
        # from LDMP.worker import AbstractWorker, StartWorker
        # from time import sleep
        #
        # class SleepWorker(AbstractWorker):
        #     def __init__(self, time):
        #         super(SleepWorker, self).__init__()
        #         self.sleep_time = time
        #
        #     def work(self):
        #         for n in range(100):
        #             if not self.killed:
        #                 sleep(self.sleep_time / float(100))
        #                 self.progress.emit(n)
        #             else:
        #                 return False
        #         return True
        #
        # # Use Trends.Earth for calculation
        # self.mode_te_prod.setChecked(True)
        #
        # # Ensure any message boxes that open are closed within 1 second
        # def close_msg_boxes():
        #     for w in QApplication.topLevelWidgets():
        #         if isinstance(w, QMessageBox):
        #             print('Closing message box')
        #             QTest.keyClick(w, Qt.Key_Enter)
        # timer = QTimer()
        # timer.timeout.connect(close_msg_boxes)
        # timer.start(1000)
        #
        # first_row = self.area_tab.area_admin_0.findText('Burundi')
        # last_row = self.area_tab.area_admin_0.findText('Portugal')
        # log('First country: {}'.format(self.area_tab.area_admin_0.itemText(first_row)))
        # log('Last country: {}'.format(self.area_tab.area_admin_0.itemText(last_row)))
        #
        # # First make sure all admin boundaries are pre-downloaded
        # for row in range(first_row, last_row):
        #     index = self.area_tab.area_admin_0.model().index(row, 0)
        #     country = self.area_tab.area_admin_0.model().data(index)
        #     adm0_a3 = self.area_tab.admin_bounds_key[country]['code']
        #     admin_polys = read_json('admin_bounds_polys_{}.json.gz'.format(adm0_a3), verify=False)
        #
        # for row in range(first_row, last_row):
        #     self.area_tab.area_admin_0.setCurrentIndex(row)
        #     index = self.area_tab.area_admin_0.model().index(row, 0)
        #     country = self.area_tab.area_admin_0.model().data(index)
        #     name = u'{}_All_Indicators_TE'.format(country)
        #     log(name)
        #     # self.options_tab.task_name.setText(name)
        #     # self.btn_calculate()
        #     
        #     # Sleep without freezing interface
        #     sleep_worker = StartWorker(SleepWorker, 'sleeping', 60)
        #     if not deg_lc_clip_worker.success:
        #         break
        #
        #######################################################################
        #######################################################################
        # End hack
        #######################################################################
        #######################################################################

    def btn_calculate(self):
        # Note that the super class has several tests in it - if they fail it
        # returns False, which would mean this function should stop execution
        # as well.
        ret = super(DlgCalculateOneStep, self).btn_calculate()
        if not ret:
            return

        if (self.year_final.date().year() - self.year_initial.date().year()) < 10:
            QtWidgets.QMessageBox.critical(None, QtWidgets.QApplication.translate("LDMP", "Error"),
                                       QtWidgets.QApplication.translate("LDMP", "Initial and final year must be at least 10 years apart."))
            return

        self.close()

        #######################################################################
        # Online

        prod_traj_year_initial = self.year_initial.date().year()
        prod_traj_year_final = self.year_final.date().year()

        prod_perf_year_initial = self.year_initial.date().year()
        prod_perf_year_final = self.year_final.date().year()

        # Have productivity state consider the last 3 years for the current 
        # period, and the years preceding those last 3 for the baseline
        prod_state_year_bl_start = self.year_initial.date().year()
        prod_state_year_bl_end = self.year_final.date().year() - 3
        prod_state_year_tg_start = prod_state_year_bl_end + 1
        prod_state_year_tg_end = prod_state_year_bl_end + 3
        assert (prod_state_year_tg_end == self.year_final.date().year())

        lc_year_initial = self.year_initial.date().year()
        lc_year_final = self.year_final.date().year()

        soc_year_initial = self.year_initial.date().year()
        soc_year_final = self.year_final.date().year()

        if self.mode_te_prod.isChecked():
            prod_mode = 'Trends.Earth productivity'
        else:
            prod_mode = 'JRC LPD'

        crosses_180th, geojsons = self.aoi.bounding_box_gee_geojson()
        payload = {'prod_mode': prod_mode,
                   'prod_traj_year_initial': prod_traj_year_initial,
                   'prod_traj_year_final': prod_traj_year_final,
                   'prod_perf_year_initial': prod_perf_year_initial,
                   'prod_perf_year_final': prod_perf_year_final,
                   'prod_state_year_bl_start': prod_state_year_bl_start,
                   'prod_state_year_bl_end': prod_state_year_bl_end,
                   'prod_state_year_tg_start': prod_state_year_tg_start,
                   'prod_state_year_tg_end': prod_state_year_tg_end,
                   'lc_year_initial': lc_year_initial,
                   'lc_year_final': lc_year_final,
                   'soc_year_initial': soc_year_initial,
                   'soc_year_final': soc_year_final,
                   'geojsons': json.dumps(geojsons),
                   'crs': self.aoi.get_crs_dst_wkt(),
                   'crosses_180th': crosses_180th,
                   'prod_traj_method': 'ndvi_trend',
                   'ndvi_gee_dataset': 'users/geflanddegradation/toolbox_datasets/ndvi_modis_2001_2016',
                   'climate_gee_dataset': None,
                   'fl': .80,
                   'trans_matrix': self.lc_define_deg_tab.trans_matrix_get(),
                   'remap_matrix': self.lc_setup_tab.dlg_esa_agg.get_agg_as_list(),
                   'task_name': self.options_tab.task_name.text(),
                   'task_notes': self.options_tab.task_notes.toPlainText()}

        resp = run_script(get_script_slug('sdg-sub-indicators'), payload)

        if resp:
            mb.pushMessage(QtWidgets.QApplication.translate("LDMP", "Submitted"),
                           QtWidgets.QApplication.translate("LDMP", "SDG sub-indicator task submitted to Google Earth Engine."),
                           level=0, duration=5)
        else:
            mb.pushMessage(QtWidgets.QApplication.translate("LDMP", "Error"),
                           QtWidgets.QApplication.translate("LDMP", "Unable to submit SDG sub-indicator task to Google Earth Engine."),
                           level=0, duration=5)


class DegradationSummaryWorkerSDG(AbstractWorker):
    def __init__(self, src_file, prod_band_nums, prod_mode, prod_out_file, 
                 lc_band_nums, soc_band_nums):
        AbstractWorker.__init__(self)

        self.src_file = src_file
        self.prod_band_nums = prod_band_nums
        self.prod_mode = prod_mode
        self.prod_out_file = prod_out_file
        # Note the first entry in the lc_band_nums, and soc_band_nums lists is
        # the degradation layer for that dataset
        self.lc_band_nums = lc_band_nums
        self.soc_band_nums = soc_band_nums

    def work(self):
        self.toggle_show_progress.emit(True)
        self.toggle_show_cancel.emit(True)

        src_ds = gdal.Open(self.src_file)

        band_lc_deg = src_ds.GetRasterBand(self.lc_band_nums[0])
        band_lc_bl = src_ds.GetRasterBand(self.lc_band_nums[1])
        band_lc_tg = src_ds.GetRasterBand(self.lc_band_nums[-1])
        band_soc_deg = src_ds.GetRasterBand(self.soc_band_nums[0])

        if self.prod_mode == 'Trends.Earth productivity':
            traj_band = src_ds.GetRasterBand(self.prod_band_nums[0])
            perf_band = src_ds.GetRasterBand(self.prod_band_nums[1])
            state_band = src_ds.GetRasterBand(self.prod_band_nums[2])
            block_sizes = traj_band.GetBlockSize()
            xsize = traj_band.XSize
            ysize = traj_band.YSize
            # Save the combined productivity indicator as well, in the second  
            # layer in the deg file
            n_out_bands = 2
        else:
            lpd_band = src_ds.GetRasterBand(self.prod_band_nums[0])
            block_sizes = band_lc_deg.GetBlockSize()
            xsize = band_lc_deg.XSize
            ysize = band_lc_deg.YSize
            n_out_bands = 1

        x_block_size = block_sizes[0]
        y_block_size = block_sizes[1]

        # Setup output file for SDG degradation indicator and combined 
        # productivity bands
        driver = gdal.GetDriverByName("GTiff")
        dst_ds_deg = driver.Create(self.prod_out_file, xsize, ysize, n_out_bands, 
                                   gdal.GDT_Int16, ['COMPRESS=LZW'])
        src_gt = src_ds.GetGeoTransform()
        dst_ds_deg.SetGeoTransform(src_gt)
        dst_srs = osr.SpatialReference()
        dst_srs.ImportFromWkt(src_ds.GetProjectionRef())
        dst_ds_deg.SetProjection(dst_srs.ExportToWkt())

        # Width of cells in longitude
        long_width = src_gt[1]
        # Set initial lat ot the top left corner latitude
        lat = src_gt[3]
        # Width of cells in latitude
        pixel_height = src_gt[5]

        trans_xtab = None
        # The 8 below is for eight classes plus no data, and the minus one is 
        # because one of the bands is a degradation layer
        soc_totals_table = [None] * (len(self.soc_band_nums) - 1)
        lc_totals_table = np.zeros((len(self.lc_band_nums) - 1, 8))
        sdg_tbl_overall = np.zeros((4, 1))
        sdg_tbl_prod = np.zeros((4, 1))
        sdg_tbl_soc = np.zeros((4, 1))
        sdg_tbl_lc = np.zeros((4, 1))

        blocks = 0
        for y in range(0, ysize, y_block_size):
            if self.killed:
                log("Processing of {} killed by user after processing {} out of {} blocks.".format(self.prod_out_file, y, ysize))
                break
            self.progress.emit(100 * float(y) / ysize)
            if y + y_block_size < ysize:
                rows = y_block_size
            else:
                rows = ysize - y
            for x in range(0, xsize, x_block_size):
                if x + x_block_size < xsize:
                    cols = x_block_size
                else:
                    cols = xsize - x

                if self.prod_mode == 'Trends.Earth productivity':
                    traj_array = traj_band.ReadAsArray(x, y, cols, rows)
                    state_array = state_band.ReadAsArray(x, y, cols, rows)
                    perf_array = perf_band.ReadAsArray(x, y, cols, rows)

                    ##############
                    # Productivity
                    
                    # Recode trajectory into deg, stable, imp. Capture trends 
                    # that are at least 95% significant.
                    #
                    # Remember that traj is coded as:
                    # -3: 99% signif decline
                    # -2: 95% signif decline
                    # -1: 90% signif decline
                    #  0: stable
                    #  1: 90% signif increase
                    #  2: 95% signif increase
                    #  3: 99% signif increase
                    traj_recode = traj_array.copy()
                    traj_recode[(traj_array >= -3) & (traj_array < -1)] = -1
                    # -1 and 1 are not signif at 95%, so stable
                    traj_recode[(traj_array >= -1) & (traj_array <= 1)] = 0
                    traj_recode[(traj_array > 1) & (traj_array <= 3)] = 1

                    # Recode state into deg, stable, imp. Note the >= -10 is so 
                    # no data isn't coded as degradation. More than two changes 
                    # in class is defined as degradation in state.
                    state_recode = state_array.copy()
                    state_recode[(state_array >= -10) & (state_array <= -2)] = -1
                    state_recode[(state_array > -2) & (state_array < 2)] = 0
                    state_recode[state_array >= 2] = 1

                    # Coding of LPD (prod5)
                    # 1: declining
                    # 2: early signs of decline
                    # 3: stable but stressed
                    # 4: stable
                    # 5: improving
                    # -32768: no data
                    prod5 = traj_recode.copy()

                    ### LPD: Declining = 1
                    prod5[traj_recode == -1] = 1
                    ### LPD: Stable = 4
                    prod5[traj_recode == 0] = 4
                    ### LPD: Improving = 5
                    prod5[traj_recode == 1] = 5

                    ##
                    # Handle state and performance.
                    
                    ### LPD: Stable due to agreement in perf and state but positive trajectory
                    prod5[(traj_recode == 1) & (state_recode == -1) & (perf_array == -1)] = 4
                    ### LPD: Stable but stressed
                    prod5[(traj_recode == 0) & (state_recode == 0) & (perf_array == -1)] = 3
                    ### LPD: Early signs of decline
                    prod5[(traj_recode == 0) & (state_recode == -1) & (perf_array == 0)] = 2

                    ##
                    # Handle NAs
                    
                    # Ensure NAs carry over to productivity indicator layer
                    prod5[traj_array == -32768] = -32768
                    prod5[perf_array == -32768] = -32768
                    prod5[state_array == -32768] = -32768

                    # Ensure masked areas carry over to productivity indicator 
                    # layer
                    prod5[traj_array == -32767] = -32767
                    prod5[perf_array == -32767] = -32767
                    prod5[state_array == -32767] = -32767

                    # Save combined productivity indicator for later visualization
                    dst_ds_deg.GetRasterBand(2).WriteArray(prod5, x, y)
                else:
                    lpd_array = lpd_band.ReadAsArray(x, y, cols, rows)
                    prod5 = lpd_array
                    # TODO: Below is temporary until missing data values are 
                    # fixed in LPD layer on GEE and missing data values are 
                    # fixed in LPD layer made by UNCCD for SIDS
                    prod5[prod5 == 0] = -32768
                    prod5[prod5 == 15] = -32768

                # Recode prod5 as stable, degraded, improved (prod3)
                prod3 = prod5.copy()
                prod3[(prod5 >= 1) & (prod5 <= 3)] = -1
                prod3[prod5 == 4] = 0
                prod3[prod5 == 5] = 1

                deg_sdg = prod3.copy()

                #############
                # Land cover
                lc_array = band_lc_deg.ReadAsArray(x, y, cols, rows)
                deg_sdg[lc_array == -1] = -1

                ##############
                # Soil carbon
                
                # Note SOC array is coded in percent change, so change of 
                # greater than 10% is improvement or decline.
                soc_array = band_soc_deg.ReadAsArray(x, y, cols, rows)
                deg_sdg[(soc_array <= -10) & (soc_array >= -100)] = -1

                #############
                # Improvement
                
                # Allow improvements by lc or soc, only where one of the other 
                # two indicators doesn't indicate a decline
                deg_sdg[(deg_sdg == 0) & (lc_array == 1)] = 1
                deg_sdg[(deg_sdg == 0) & (soc_array >= 10) & (soc_array <= 100)] = 1

                ##############
                # Missing data
                
                # Ensure all NAs are carried over - note this was already done 
                # for the productivity layer above but need to do it again in 
                # case values from another layer overwrote those missing value 
                # indicators.
                
                # No data
                deg_sdg[prod3 == -32768] = -32768
                deg_sdg[lc_array == -32768] = -32768
                deg_sdg[soc_array == -32768] = -32768

                deg_sdg[prod3 == -32767] = -32767
                deg_sdg[lc_array == -32767] = -32767
                deg_sdg[soc_array == -32767] = -32767

                dst_ds_deg.GetRasterBand(1).WriteArray(deg_sdg, x, y)

                ###########################################################
                # Start areal calculations
                ###########################################################

                # Caculate cell area for each horizontal line
                cell_areas = np.array([calc_cell_area(lat + pixel_height*n, lat + pixel_height*(n + 1), long_width) for n in range(rows)])
                cell_areas.shape = (cell_areas.size, 1)
                # Make an array of the same size as the input arrays containing 
                # the area of each cell (which is identical for all cells ina 
                # given row - cell areas only vary among rows)
                cell_areas_array = np.repeat(cell_areas, cols, axis=1)

                a_lc_bl = band_lc_bl.ReadAsArray(x, y, cols, rows)
                a_lc_tg = band_lc_tg.ReadAsArray(x, y, cols, rows)

                ###########################################################
                # Tabulate SDG 15.3.1 indicator
                
                # Mask water areas
                deg_sdg[a_lc_tg == 7] = -32767
                sdg_tbl_overall[0] = sdg_tbl_overall[0] + np.sum((deg_sdg == 1) * cell_areas_array)
                sdg_tbl_overall[1] = sdg_tbl_overall[1] + np.sum((deg_sdg == 0) * cell_areas_array)
                sdg_tbl_overall[2] = sdg_tbl_overall[2] + np.sum((deg_sdg == -1) * cell_areas_array)
                sdg_tbl_overall[3] = sdg_tbl_overall[3] + np.sum((deg_sdg == -32768) * cell_areas_array)

                ###########################################################
                # Calculate transition crosstabs for productivity indicator, 
                # and SOC totals
                a_trans_bl_tg = a_lc_bl*10 + a_lc_tg
                a_trans_bl_tg[np.logical_or(a_lc_bl < 1, a_lc_tg < 1)] <- -32768

                # Mask water areas
                prod3[a_lc_tg == 7] = -32767
                sdg_tbl_prod[0] = sdg_tbl_prod[0] + np.sum((prod3 == 1) * cell_areas_array)
                sdg_tbl_prod[1] = sdg_tbl_prod[1] + np.sum((prod3 == 0) * cell_areas_array)
                sdg_tbl_prod[2] = sdg_tbl_prod[2] + np.sum((prod3 == -1) * cell_areas_array)
                sdg_tbl_prod[3] = sdg_tbl_prod[3] + np.sum((prod3 == -32768) * cell_areas_array)

                # Calculate the xtab and soc totals row by row over the y rows 
                # since each one needs to be weighted by the pixel area for 
                # that row
                for n in range(y_block_size):
                    cell_area = cell_areas[n]
                    this_trans_xtab = xtab(prod5[n, :], a_trans_bl_tg[n, :])
                    # Don't use this_trans_xtab if it is empty (could happen if 
                    # take a crosstab where all of the values are nan's)
                    if this_trans_xtab[0][0].size != 0:
                        this_trans_xtab[1] = this_trans_xtab[1] * cell_area
                        if trans_xtab == None:
                            trans_xtab = this_trans_xtab
                        else:
                            trans_xtab = merge_xtabs(trans_xtab, this_trans_xtab)

                    # Calculate SOC totals). Note final units of soc_totals 
                    # tables are tons C (summed over the total area of each 
                    # class). Start at one because the first soc band is the 
                    # degradation layer.
                    for i in range(1, len(self.soc_band_nums)):
                        band_soc = src_ds.GetRasterBand(self.soc_band_nums[i])
                        a_soc = band_soc.ReadAsArray(x, y, cols, rows)
                        # Convert soilgrids data from per ha to per meter since 
                        # cell_area is in meters
                        a_soc = a_soc.astype(np.float32) / (100 * 100)
                        soc_totals_table[i - 1] = calc_total_table(a_trans_bl_tg[n, :],
                                                                   a_soc[n, :], 
                                                                   soc_totals_table[i - 1], 
                                                                   cell_area)
                band_soc_bl = src_ds.GetRasterBand(self.soc_band_nums[1])
                a_soc_bl = band_soc_bl.ReadAsArray(x, y, cols, rows)
                a_soc_bl = a_soc_bl.astype(np.float32) / (100 * 100)
                band_soc_tg = src_ds.GetRasterBand(self.soc_band_nums[-1])
                a_soc_tg = band_soc_tg.ReadAsArray(x, y, cols, rows)
                # Save masked and nodata values before they are obliterated by 
                # the units change below (the 100 * 100)
                a_soc_tg_masked = a_soc_tg == -32767
                a_soc_tg_nodata = a_soc_tg == -32768
                a_soc_tg = a_soc_tg.astype(np.float32) / (100 * 100)

                a_soc_frac_chg = a_soc_tg / a_soc_bl
                # Degradation in terms of SOC is defined as a decline of more 
                # than 10% (and improving increase greater than 10%)
                a_deg_soc = a_soc_frac_chg.copy()
                a_deg_soc[(a_soc_frac_chg >= 0) & (a_soc_frac_chg <= .9)] = -1
                a_deg_soc[(a_soc_frac_chg > .9) & (a_soc_frac_chg < 1.1)] = 0
                a_deg_soc[a_soc_frac_chg >= 1.1] = 1
                # Carry over areas that were originally masked or no data
                a_deg_soc[a_soc_tg_masked] = -32767 # Masked areas
                a_deg_soc[a_soc_tg_nodata] = -32768 # No data
                # Mask water areas
                a_deg_soc[a_lc_tg == 7] = -32767
                sdg_tbl_soc[0] = sdg_tbl_soc[0] + np.sum((a_deg_soc == 1) * cell_areas_array)
                sdg_tbl_soc[1] = sdg_tbl_soc[1] + np.sum((a_deg_soc == 0) * cell_areas_array)
                sdg_tbl_soc[2] = sdg_tbl_soc[2] + np.sum((a_deg_soc == -1) * cell_areas_array)
                sdg_tbl_soc[3] = sdg_tbl_soc[3] + np.sum((a_deg_soc == -32768) * cell_areas_array)

                # Start at one because remember the first lc band is the 
                # degradation layer
                for i in range(1, len(self.lc_band_nums)):
                    band_lc = src_ds.GetRasterBand(self.lc_band_nums[i])
                    a_lc = band_lc.ReadAsArray(x, y, cols, rows)
                    lc_totals_table[i - 1] = np.add([np.sum((a_lc == c) * cell_areas_array) for c in [1, 2, 3, 4, 5, 6, 7, -32768]], lc_totals_table[i - 1])

                a_deg_lc = band_lc_deg.ReadAsArray(x, y, cols, rows)
                a_deg_lc[a_lc_tg == 7] = -32767
                sdg_tbl_lc[0] = sdg_tbl_lc[0] + np.sum((a_deg_lc == 1) * cell_areas_array)
                sdg_tbl_lc[1] = sdg_tbl_lc[1] + np.sum((a_deg_lc == 0) * cell_areas_array)
                sdg_tbl_lc[2] = sdg_tbl_lc[2] + np.sum((a_deg_lc == -1) * cell_areas_array)
                sdg_tbl_lc[3] = sdg_tbl_lc[3] + np.sum((a_deg_lc == -32768) * cell_areas_array)

                blocks += 1
            lat += pixel_height * rows
        self.progress.emit(100)

        if self.killed:
            del dst_ds_deg
            os.remove(self.prod_out_file)
            return None
        else:
            # Convert all area tables from meters into square kilometers
            trans_xtab[1] = trans_xtab[1] * 1e-6
            sdg_tbl_overall = sdg_tbl_overall * 1e-6
            sdg_tbl_prod = sdg_tbl_prod * 1e-6
            sdg_tbl_soc = sdg_tbl_soc * 1e-6
            sdg_tbl_lc = sdg_tbl_lc * 1e-6
            lc_totals_table = lc_totals_table * 1e-6

            return list((soc_totals_table, lc_totals_table, trans_xtab, 
                         sdg_tbl_overall, sdg_tbl_prod, sdg_tbl_soc, sdg_tbl_lc))


class DlgCalculateSummaryTableAdmin(DlgCalculateBase, Ui_DlgCalculateSummaryTableAdmin):
    def __init__(self, parent=None):
        super(DlgCalculateSummaryTableAdmin, self).__init__(parent)

        self.setupUi(self)

        self.mode_lpd_jrc.toggled.connect(self.mode_lpd_jrc_toggled)
        self.mode_lpd_jrc_toggled()

        self.browse_output_file_layer.clicked.connect(self.select_output_file_layer)
        self.browse_output_file_table.clicked.connect(self.select_output_file_table)

    def mode_lpd_jrc_toggled(self):
        if self.mode_lpd_jrc.isChecked():
            self.combo_layer_lpd.setEnabled(True)
            self.combo_layer_traj.setEnabled(False)
            self.combo_layer_traj_label.setEnabled(False)
            self.combo_layer_perf.setEnabled(False)
            self.combo_layer_perf_label.setEnabled(False)
            self.combo_layer_state.setEnabled(False)
            self.combo_layer_state_label.setEnabled(False)
        else:
            self.combo_layer_lpd.setEnabled(False)
            self.combo_layer_traj.setEnabled(True)
            self.combo_layer_traj_label.setEnabled(True)
            self.combo_layer_perf.setEnabled(True)
            self.combo_layer_perf_label.setEnabled(True)
            self.combo_layer_state.setEnabled(True)
            self.combo_layer_state_label.setEnabled(True)

    def showEvent(self, event):
        super(DlgCalculateSummaryTableAdmin, self).showEvent(event)

        self.combo_layer_lpd.populate()
        self.combo_layer_traj.populate()
        self.combo_layer_perf.populate()
        self.combo_layer_state.populate()
        self.combo_layer_lc.populate()
        self.combo_layer_soc.populate()

    def select_output_file_layer(self):
        f = QtWidgets.QFileDialog.getSaveFileName(self,
                                              self.tr('Choose a filename for the output file'),
                                              QSettings().value("LDMP/output_dir", None),
                                              self.tr('Filename (*.json)'))
        if f:
            if os.access(os.path.dirname(f), os.W_OK):
                QSettings().setValue("LDMP/output_dir", os.path.dirname(f))
                self.output_file_layer.setText(f)
            else:
                QtWidgets.QMessageBox.critical(None, self.tr("Error"),
                                           self.tr(u"Cannot write to {}. Choose a different file.".format(f), None))

    def select_output_file_table(self):
        f = QtWidgets.QFileDialog.getSaveFileName(self,
                                              self.tr('Choose a filename for the summary table'),
                                              QSettings().value("LDMP/output_dir", None),
                                              self.tr('Summary table file (*.xlsx)'))
        if f:
            if os.access(os.path.dirname(f), os.W_OK):
                QSettings().setValue("LDMP/output_dir", os.path.dirname(f))
                self.output_file_table.setText(f)
            else:
                QtWidgets.QMessageBox.critical(None, self.tr("Error"),
                                           self.tr(u"Cannot write to {}. Choose a different file.".format(f), None))

    def btn_calculate(self):
        ######################################################################
        # Check that all needed output files are selected
        if not self.output_file_layer.text():
            QtWidgets.QMessageBox.information(None, self.tr("Error"),
                                          self.tr("Choose an output file for the indicator layer."), None)
            return

        if not self.output_file_table.text():
            QtWidgets.QMessageBox.information(None, self.tr("Error"),
                                          self.tr("Choose an output file for the summary table."), None)
            return

        # Note that the super class has several tests in it - if they fail it
        # returns False, which would mean this function should stop execution
        # as well.
        ret = super(DlgCalculateSummaryTableAdmin, self).btn_calculate()
        if not ret:
            return

        if self.mode_te_prod.isChecked():
            prod_mode = 'Trends.Earth productivity'
        else:
            prod_mode = 'JRC LPD'


        ######################################################################
        # Check that all needed input layers are selected
        if prod_mode == 'Trends.Earth productivity':
            if len(self.combo_layer_traj.layer_list) == 0:
                QtWidgets.QMessageBox.critical(None, self.tr("Error"),
                                           self.tr("You must add a productivity trajectory indicator layer to your map before you can use the SDG calculation tool."), None)
                return
            if len(self.combo_layer_state.layer_list) == 0:
                QtWidgets.QMessageBox.critical(None, self.tr("Error"),
                                           self.tr("You must add a productivity state indicator layer to your map before you can use the SDG calculation tool."), None)
                return
            if len(self.combo_layer_perf.layer_list) == 0:
                QtWidgets.QMessageBox.critical(None, self.tr("Error"),
                                           self.tr("You must add a productivity performance indicator layer to your map before you can use the SDG calculation tool."), None)
                return

        else:
            if len(self.combo_layer_lpd.layer_list) == 0:
                QtWidgets.QMessageBox.critical(None, self.tr("Error"),
                                           self.tr("You must add a land productivity dynamics indicator layer to your map before you can use the SDG calculation tool."), None)
                return

        if len(self.combo_layer_lc.layer_list) == 0:
            QtWidgets.QMessageBox.critical(None, self.tr("Error"),
                                       self.tr("You must add a land cover indicator layer to your map before you can use the SDG calculation tool."), None)
            return

        if len(self.combo_layer_soc.layer_list) == 0:
            QtWidgets.QMessageBox.critical(None, self.tr("Error"),
                                       self.tr("You must add a soil organic carbon indicator layer to your map before you can use the SDG calculation tool."), None)
            return

        #######################################################################
        # Check that the layers cover the full extent needed
        if prod_mode == 'Trends.Earth productivity':
            if self.aoi.calc_frac_overlap(QgsGeometry.fromRect(self.combo_layer_traj.get_layer().extent())) < .99:
                QtWidgets.QMessageBox.critical(None, self.tr("Error"),
                                           self.tr("Area of interest is not entirely within the trajectory layer."), None)
                return
            if self.aoi.calc_frac_overlap(QgsGeometry.fromRect(self.combo_layer_perf.get_layer().extent())) < .99:
                QtWidgets.QMessageBox.critical(None, self.tr("Error"),
                                           self.tr("Area of interest is not entirely within the performance layer."), None)
                return
            if self.aoi.calc_frac_overlap(QgsGeometry.fromRect(self.combo_layer_state.get_layer().extent())) < .99:
                QtWidgets.QMessageBox.critical(None, self.tr("Error"),
                                           self.tr("Area of interest is not entirely within the state layer."), None)
                return
        else:
            if self.aoi.calc_frac_overlap(QgsGeometry.fromRect(self.combo_layer_lpd.get_layer().extent())) < .99:
                QtWidgets.QMessageBox.critical(None, self.tr("Error"),
                                           self.tr("Area of interest is not entirely within the land productivity dynamics layer."), None)
                return

        if self.aoi.calc_frac_overlap(QgsGeometry.fromRect(self.combo_layer_lc.get_layer().extent())) < .99:
            QtWidgets.QMessageBox.critical(None, self.tr("Error"),
                                       self.tr("Area of interest is not entirely within the land cover layer."), None)
            return
        if self.aoi.calc_frac_overlap(QgsGeometry.fromRect(self.combo_layer_soc.get_layer().extent())) < .99:
            QtWidgets.QMessageBox.critical(None, self.tr("Error"),
                                       self.tr("Area of interest is not entirely within the soil organic carbon layer."), None)
            return

        #######################################################################
        # Check that all of the productivity layers have the same resolution 
        # and CRS
        def res(layer):
            return (round(layer.rasterUnitsPerPixelX(), 10), round(layer.rasterUnitsPerPixelY(), 10))

        if prod_mode == 'Trends.Earth productivity':
            if res(self.combo_layer_traj.get_layer()) != res(self.combo_layer_state.get_layer()):
                QtWidgets.QMessageBox.critical(None, self.tr("Error"),
                                           self.tr("Resolutions of trajectory layer and state layer do not match."), None)
                return
            if res(self.combo_layer_traj.get_layer()) != res(self.combo_layer_perf.get_layer()):
                QtWidgets.QMessageBox.critical(None, self.tr("Error"),
                                           self.tr("Resolutions of trajectory layer and performance layer do not match."), None)
                return

            if self.combo_layer_traj.get_layer().crs() != self.combo_layer_state.get_layer().crs():
                QtWidgets.QMessageBox.critical(None, self.tr("Error"),
                                           self.tr("Coordinate systems of trajectory layer and state layer do not match."), None)
                return
            if self.combo_layer_traj.get_layer().crs() != self.combo_layer_perf.get_layer().crs():
                QtWidgets.QMessageBox.critical(None, self.tr("Error"),
                                           self.tr("Coordinate systems of trajectory layer and performance layer do not match."), None)
                return

        self.close()

        #######################################################################
        # Load all datasets to VRTs (to select only the needed bands)
        if prod_mode == 'Trends.Earth productivity':
            traj_vrt = self.combo_layer_traj.get_vrt()
            perf_vrt = self.combo_layer_perf.get_vrt()
            state_vrt = self.combo_layer_state.get_vrt()
        else:
            lpd_vrt = self.combo_layer_lpd.get_vrt()

        #######################################################################
        # Select baseline and target land cover and SOC layers based on chosen
        # degradation layers for these datasets
        lc_band_infos = get_band_infos(self.combo_layer_lc.get_data_file())
        lc_annual_band_indices = [i for i, bi in enumerate(lc_band_infos) if bi['name'] == 'Land cover (7 class)']
        lc_annual_band_indices.sort(key=lambda i: lc_band_infos[i]['metadata']['year'])
        lc_years = [bi['metadata']['year'] for bi in lc_band_infos if bi['name'] == 'Land cover (7 class)']

        soc_band_infos = get_band_infos(self.combo_layer_soc.get_data_file())
        soc_annual_band_indices = [i for i, bi in enumerate(soc_band_infos) if bi['name'] == 'Soil organic carbon']
        soc_annual_band_indices.sort(key=lambda i: soc_band_infos[i]['metadata']['year'])
        soc_years = [bi['metadata']['year'] for bi in soc_band_infos if bi['name'] == 'Soil organic carbon']

        # Make the LC degradation file first in the list
        lc_deg_f = self.combo_layer_lc.get_vrt()
        lc_files = [lc_deg_f]
        for i in lc_annual_band_indices:
            f = tempfile.NamedTemporaryFile(suffix='.vrt').name
            # Add once since band numbers don't start at zero
            gdal.BuildVRT(f,
                          self.combo_layer_lc.get_data_file(),
                          bandList=[i + 1])
            lc_files.append(f)

        # Make the SOC degradation file first in the list
        soc_deg_f = self.combo_layer_soc.get_vrt()
        soc_files = [soc_deg_f]
        for i in soc_annual_band_indices:
            f = tempfile.NamedTemporaryFile(suffix='.vrt').name
            # Add once since band numbers don't start at zero
            gdal.BuildVRT(f,
                          self.combo_layer_soc.get_data_file(),
                          bandList=[i + 1])
            soc_files.append(f)

        in_files = list(lc_files)
        in_files.extend(soc_files)
        lc_band_nums = np.arange(len(lc_files)) + 1
        soc_band_nums = np.arange(len(soc_files)) + 1 + lc_band_nums.max()

        # Remember the first value is an indication of whether dataset is 
        # wrapped across 180th meridian
        wkts = self.aoi.meridian_split('layer', 'wkt', warn=False)[1]
        if prod_mode == 'Trends.Earth productivity':
            bbs = self.aoi.get_aligned_output_bounds(traj_vrt)
        else:
            bbs = self.aoi.get_aligned_output_bounds(lpd_vrt)

        output_sdg_tifs = []
        output_sdg_json = self.output_file_layer.text()
        for n in range(len(wkts)):
            # Compute the pixel-aligned bounding box (slightly larger than 
            # aoi). Use this instead of croptocutline in gdal.Warp in order to 
            # keep the pixels aligned with the chosen productivity layer.
        
            # Combines SDG 15.3.1 input raster into a VRT and crop to the AOI
            indic_vrt = tempfile.NamedTemporaryFile(suffix='.vrt').name
            log(u'Saving indicator VRT to: {}'.format(indic_vrt))
            # The plus one is because band numbers start at 1, not zero
            if prod_mode == 'Trends.Earth productivity':
                in_files.extend([traj_vrt, perf_vrt, state_vrt])
                gdal.BuildVRT(indic_vrt,
                              in_files,
                              outputBounds=bbs[n],
                              resolution='highest',
                              resampleAlg=gdal.GRA_NearestNeighbour,
                              separate=True)
                prod_band_nums = np.arange(3) + 1 + soc_band_nums.max()
            else:
                in_files.append(lpd_vrt)
                gdal.BuildVRT(indic_vrt,
                              in_files,
                              outputBounds=bbs[n],
                              resolution='highest',
                              resampleAlg=gdal.GRA_NearestNeighbour,
                              separate=True)
                prod_band_nums = [max(soc_band_nums) + 1]

            masked_vrt = tempfile.NamedTemporaryFile(suffix='.tif').name
            log(u'Saving deg/lc clipped file to {}'.format(masked_vrt))
            deg_lc_clip_worker = StartWorker(ClipWorker, 'masking layers (part {} of {})'.format(n + 1, len(wkts)), 
                                             indic_vrt, masked_vrt, 
                                             json.loads(QgsGeometry.fromWkt(wkts[n]).exportToGeoJSON()),
                                             bbs[n])
            if not deg_lc_clip_worker.success:
                QtWidgets.QMessageBox.critical(None, self.tr("Error"),
                                           self.tr("Error masking SDG 15.3.1 input layers."), None)
                return

            ######################################################################
            #  Calculate SDG 15.3.1 layers
            log('Calculating summary table...')
            if len(wkts) > 1:
                output_sdg_tif = os.path.splitext(output_sdg_json)[0] + '_{}.tif'.format(n)
            else:
                output_sdg_tif = os.path.splitext(output_sdg_json)[0] + '.tif'

            output_sdg_tifs.append(output_sdg_tif)
            deg_worker = StartWorker(DegradationSummaryWorkerSDG,
                                    'calculating summary table (part {} of {})'.format(n + 1, len(wkts)),
                                     masked_vrt,
                                     prod_band_nums,
                                     prod_mode, 
                                     output_sdg_tif,
                                     lc_band_nums, 
                                     soc_band_nums)
            if not deg_worker.success:
                QtWidgets.QMessageBox.critical(None, self.tr("Error"),
                                           self.tr("Error calculating SDG 15.3.1 summary table."), None)
                return
            else:
                if n == 0:
                    soc_totals, \
                            lc_totals, \
                            trans_prod_xtab, \
                            sdg_tbl_overall, \
                            sdg_tbl_prod, \
                            sdg_tbl_soc, \
                            sdg_tbl_lc = deg_worker.get_return()
                else:
                    this_soc_totals, \
                            this_lc_totals, \
                            this_trans_prod_xtab, \
                            this_sdg_tbl_overall, \
                            this_sdg_tbl_prod, \
                            this_sdg_tbl_soc, \
                            this_sdg_tbl_lc = deg_worker.get_return()

                    log('soc_totals: {}'.format(soc_totals))
                    log('this_soc_totals: {}'.format(this_soc_totals))
                    for n in range(len(soc_totals)):
                        soc_totals[n] = merge_area_tables(soc_totals[n], this_soc_totals[n])
                    log('soc_totals after addition: {}'.format(this_soc_totals))
                    lc_totals = lc_totals + this_lc_totals
                    if this_trans_prod_xtab[0][0].size != 0:
                        trans_prod_xtab = merge_xtabs(trans_prod_xtab, this_trans_prod_xtab)
                    sdg_tbl_overall = sdg_tbl_overall + this_sdg_tbl_overall
                    sdg_tbl_prod = sdg_tbl_prod + this_sdg_tbl_prod
                    sdg_tbl_soc = sdg_tbl_soc + this_sdg_tbl_soc
                    sdg_tbl_lc = sdg_tbl_lc + this_sdg_tbl_lc

        make_summary_table(soc_totals, lc_totals, trans_prod_xtab, 
                           sdg_tbl_overall, sdg_tbl_prod, sdg_tbl_soc, 
                           sdg_tbl_lc, lc_years, soc_years,
                           self.output_file_table.text())


        # Add the SDG layers to the map
        output_sdg_bandinfos = [BandInfo("SDG 15.3.1 Indicator")]
        if prod_mode == 'Trends.Earth productivity':
            output_sdg_bandinfos.append(BandInfo("SDG 15.3.1 Productivity Indicator"))
        
        if len(output_sdg_tifs) == 1:
            output_file = output_sdg_tifs[0]
        else:
            output_file = os.path.splitext(output_sdg_json)[0] + '.vrt'
            gdal.BuildVRT(output_file, output_sdg_tifs)
        create_local_json_metadata(output_sdg_json, output_file, 
                output_sdg_bandinfos, metadata={'task_name': self.options_tab.task_name.text(),
                                                'task_notes': self.options_tab.task_notes.toPlainText()})
        schema = BandInfoSchema()
        add_layer(output_file, 1, schema.dump(output_sdg_bandinfos[0]))
        if prod_mode == 'Trends.Earth productivity':
            add_layer(output_file, 2, schema.dump(output_sdg_bandinfos[1]))


def get_lc_area(table, code):
    ind = np.where(table[0] == code)[0]
    if ind.size == 0:
        return 0
    else:
        return float(table[1][ind])


def get_lpd_table(table,
                  lc_classes=list(range(1, 6 + 1)), # Don't include water bodies in the table
                  lpd_classes=[1, 2, 3, 4, 5, -32768]):
    out = np.zeros((len(lc_classes), len(lpd_classes)))
    for lc_class_num in range(len(lc_classes)):
        for prod_num in range(len(lpd_classes)):
            transition = int('{}{}'.format(lc_classes[lc_class_num], lc_classes[lc_class_num]))
            out[lc_class_num, prod_num] = get_xtab_area(table, lpd_classes[prod_num], transition)
    return out


def get_prod_table(table, prod_class, classes=list(range(1, 7 + 1))):
    out = np.zeros((len(classes), len(classes)))
    for bl_class in range(len(classes)):
        for tg_class in range(len(classes)):
            transition = int('{}{}'.format(classes[bl_class], classes[tg_class]))
            out[bl_class, tg_class] = get_xtab_area(table, prod_class, transition)
    return out



# Note classes for this function go from 1-6 to exclude water from the SOC 
# totals
def write_soc_stock_change_table(sheet, first_row, first_col, soc_bl_totals, 
                                 soc_tg_totals, classes=list(range(1, 6 + 1))):
    for row in range(len(classes)):
        for col in range(len(classes)):
            cell = sheet.cell(row=row + first_row, column=col + first_col)
            transition = int('{}{}'.format(classes[row], classes[col]))
            bl_soc = get_soc_total(soc_bl_totals, transition)
            tg_soc = get_soc_total(soc_tg_totals, transition)
            try:
                cell.value = (tg_soc - bl_soc) / bl_soc
            except ZeroDivisionError:
                cell.value = ''
    

# Note classes for this function go from 1-6 to exclude water from the SOC 
# totals
def get_soc_total_by_class(trans_prod_xtab, soc_totals, classes=list(range(1, 6 + 1))):
    out = np.zeros((len(classes), 1))
    for row in range(len(classes)):
        area = 0
        soc = 0
        # Need to sum up the total soc across the pixels and then divide by 
        # total area
        for n in range(len(classes)):
            trans = int('{}{}'.format(classes[row], classes[n]))
            area += get_xtab_area(trans_prod_xtab, None, trans)
            soc += get_soc_total(soc_totals, trans)

        # Note areas are in sq km. Need to convert to ha
        if soc != 0 and area != 0:
            out[row][0] = soc / (area * 100)
        else:
            out[row][0]
    return out


def get_lc_table(table, classes=list(range(1, 7 + 1))):
    out = np.zeros((len(classes), len(classes)))
    for bl_class in range(len(classes)):
        for tg_class in range(len(classes)):
            transition = int('{}{}'.format(classes[bl_class], classes[tg_class]))
            out[bl_class, tg_class] = get_xtab_area(table, None, transition)
    return out


def get_soc_total(soc_table, transition):
    ind = np.where(soc_table[0] == transition)[0]
    if ind.size == 0:
        return 0
    else:
        return float(soc_table[1][ind])


def make_summary_table(soc_totals, lc_totals, trans_prod_xtab, sdg_tbl_overall, 
                       sdg_tbl_prod, sdg_tbl_soc, sdg_tbl_lc, lc_years, 
                       soc_years, out_file):
    def tr(s):
        return QtWidgets.QApplication.translate("LDMP", s)

    wb = openpyxl.load_workbook(os.path.join(os.path.dirname(os.path.realpath(__file__)), 'data', 'SummaryTable.xlsx'))

    ##########################################################################
    # SDG table
    ws_sdg = wb.get_sheet_by_name('SDG 15.3.1')
    write_table_to_sheet(ws_sdg, sdg_tbl_overall, 6, 6)

    ##########################################################################
    # Productivity tables
    ws_prod = wb.get_sheet_by_name('Productivity')
    write_table_to_sheet(ws_prod, sdg_tbl_prod, 6, 6)

    write_table_to_sheet(ws_prod, get_prod_table(trans_prod_xtab, 5), 16, 3)
    write_table_to_sheet(ws_prod, get_prod_table(trans_prod_xtab, 4), 28, 3)
    write_table_to_sheet(ws_prod, get_prod_table(trans_prod_xtab, 3), 40, 3)
    write_table_to_sheet(ws_prod, get_prod_table(trans_prod_xtab, 2), 52, 3)
    write_table_to_sheet(ws_prod, get_prod_table(trans_prod_xtab, 1), 64, 3)
    write_table_to_sheet(ws_prod, get_prod_table(trans_prod_xtab, -32768), 76, 3)

    ##########################################################################
    # Soil organic carbon tables
    ws_soc = wb.get_sheet_by_name('Soil organic carbon')
    write_table_to_sheet(ws_soc, sdg_tbl_soc, 6, 6)

    # First write baseline
    write_table_to_sheet(ws_soc, get_soc_total_by_class(trans_prod_xtab, soc_totals[0]), 16, 3)
    # Now write target
    write_table_to_sheet(ws_soc, get_soc_total_by_class(trans_prod_xtab, soc_totals[-1]), 16, 4)

    # Write table of baseline areas
    lc_trans_table_no_water = get_lc_table(trans_prod_xtab, classes=np.arange(1, 6 + 1))
    write_table_to_sheet(ws_soc, np.reshape(np.sum(lc_trans_table_no_water, 1), (-1, 1)), 16, 5)
    # Write table of target areas
    write_table_to_sheet(ws_soc, np.reshape(np.sum(lc_trans_table_no_water, 0), (-1, 1)), 16, 6)
    
    # write_soc_stock_change_table has its own writing function as it needs to write a 
    # mix of numbers and strings
    write_soc_stock_change_table(ws_soc, 27, 3, soc_totals[0], soc_totals[-1])

    ##########################################################################
    # Land cover tables
    ws_lc = wb.get_sheet_by_name('Land cover')
    write_table_to_sheet(ws_lc, sdg_tbl_lc, 6, 6)

    write_table_to_sheet(ws_lc, get_lc_table(trans_prod_xtab), 26, 3)

    ##########################################################################
    # UNCCD tables
    ws_unccd = wb.get_sheet_by_name('UNCCD Reporting')

    for i in range(len(lc_years)):
        # Water bodies
        cell = ws_unccd.cell(5 + i, 4)
        cell.value = lc_totals[i][6]
        # Other classes
        write_row_to_sheet(ws_unccd, np.append(lc_years[i], lc_totals[i][0:6]), 24 + i, 2)

    write_table_to_sheet(ws_unccd, get_lpd_table(trans_prod_xtab), 54, 3)

    for i in range(len(soc_years)):
        write_row_to_sheet(ws_unccd, np.append(soc_years[i], get_soc_total_by_class(trans_prod_xtab, soc_totals[i])), 64 + i, 2)

    try:
        ws_sdg_logo = Image(os.path.join(os.path.dirname(__file__), 'data', 'trends_earth_logo_bl_300width.png'))
        ws_sdg.add_image(ws_sdg_logo, 'H1')
        ws_prod_logo = Image(os.path.join(os.path.dirname(__file__), 'data', 'trends_earth_logo_bl_300width.png'))
        ws_prod.add_image(ws_prod_logo, 'H1')
        ws_soc_logo = Image(os.path.join(os.path.dirname(__file__), 'data', 'trends_earth_logo_bl_300width.png'))
        ws_soc.add_image(ws_soc_logo, 'H1')
        ws_lc_logo = Image(os.path.join(os.path.dirname(__file__), 'data', 'trends_earth_logo_bl_300width.png'))
        ws_lc.add_image(ws_lc_logo, 'H1')
        ws_unccd_logo = Image(os.path.join(os.path.dirname(__file__), 'data', 'trends_earth_logo_bl_300width.png'))
        ws_unccd.add_image(ws_unccd_logo, 'G1')
    except ImportError:
        # add_image will fail on computers without PIL installed (this will be 
        # an issue on some Macs, likely others). it is only used here to add 
        # our logo, so no big deal.
        pass

    try:
        wb.save(out_file)
        log(u'Indicator table saved to {}'.format(out_file))
        QtWidgets.QMessageBox.information(None, QtWidgets.QApplication.translate("LDMP", "Success"),
                                      QtWidgets.QApplication.translate("LDMP", u'Indicator table saved to {}'.format(out_file)))

    except IOError:
        log(u'Error saving {}'.format(out_file))
        QtWidgets.QMessageBox.critical(None, QtWidgets.QApplication.translate("LDMP", "Error"),
                                   QtWidgets.QApplication.translate("LDMP", u"Error saving output table - check that {} is accessible and not already open.".format(out_file)), None)
